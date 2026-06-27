"""
Appends an expense row to the existing March Partners Entity Expenses tracker.

Sheet structure (per tab):
  Row 1    : empty
  Row 2    : "ENTITY EXPENSES"
  Row 3    : Full entity name
  Row 4    : Headers:
             [blank | # | Item | Category | Amount (SGD) w GST | Date |
              Payer | Used as offset vs capital call? | Comment]
  Row 5+   : Data rows

Field mapping from bot:
  Purpose  → Item  (col C)
  Category → left blank (col D)
  Amount   → Amount (SGD) w GST (col E)
  Date     → Date (col F)
  Payer    → Payer (col G)
  Offset   → always "N/A" (col H)
  Filename → Comment (col I)
"""

import io
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Side

logger = logging.getLogger(__name__)

# Column indices (1-based, matching actual spreadsheet layout)
COL_BLANK   = 1
COL_NUM     = 2   # # — auto-increment formula
COL_ITEM    = 3   # Item / Purpose
COL_CAT     = 4   # Category (left blank)
COL_AMOUNT  = 5   # Amount (SGD) w GST
COL_DATE    = 6   # Date
COL_PAYER   = 7   # Payer
COL_OFFSET  = 8   # Used as offset vs capital call?
COL_COMMENT = 9   # Comment (receipt filename stored here)

HEADER_ROW = 4
DATA_START = 5


def _last_data_row(ws) -> int:
    """Row index of the last row containing any data at or below DATA_START."""
    last = HEADER_ROW
    for row in ws.iter_rows(min_row=DATA_START):
        if any(cell.value is not None for cell in row):
            last = row[0].row
    return last


def _copy_row_style(ws, src_row: int, dst_row: int) -> None:
    """Copy cell styles from one row to another (best-effort)."""
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font          = src.font.copy()
            dst.fill          = src.fill.copy()
            dst.border        = src.border.copy()
            dst.alignment     = src.alignment.copy()
            dst.number_format = src.number_format


def add_expense(
    workbook_bytes: bytes,
    sheet_name: str,
    purpose: str,
    amount: float | None,
    expense_date: datetime,
    payer: str,
    receipt_filename: str,
) -> bytes:
    """
    Append one expense row to `sheet_name` and return the updated workbook bytes.

    - purpose        → Item column
    - amount         → Amount (SGD) w GST column
    - expense_date   → Date column
    - payer          → Payer column
    - offset         → always "N/A"
    - receipt_filename → Comment column
    """
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes))

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws       = wb[sheet_name]
    last_row = _last_data_row(ws)
    new_row  = last_row + 1

    logger.info("Appending to '%s' at row %d", sheet_name, new_row)

    # Carry forward styling from the previous data row
    if last_row >= DATA_START:
        _copy_row_style(ws, last_row, new_row)

    # # column — auto-increment formula (literal 1 for very first entry)
    if last_row == HEADER_ROW:
        ws.cell(new_row, COL_NUM).value = 1
    else:
        ws.cell(new_row, COL_NUM).value = f"=B{last_row}+1"

    # Data columns
    ws.cell(new_row, COL_ITEM).value    = purpose
    ws.cell(new_row, COL_CAT).value     = None              # Category — intentionally blank

    amount_cell = ws.cell(new_row, COL_AMOUNT)
    amount_cell.value = amount
    if amount is not None:
        amount_cell.number_format = '#,##0.00'

    date_cell = ws.cell(new_row, COL_DATE)
    date_cell.value         = expense_date
    date_cell.number_format = "D MMM YYYY"

    ws.cell(new_row, COL_PAYER).value   = payer
    ws.cell(new